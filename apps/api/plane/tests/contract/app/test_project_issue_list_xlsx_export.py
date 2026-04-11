# Copyright (c) 2023-present Plane Software, Inc. and contributors
# SPDX-License-Identifier: AGPL-3.0-only
# See the LICENSE file for details.

from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook
from rest_framework import status

from plane.db.models import (
    Issue,
    IssueAssignee,
    IssueLabel,
    IssueWorkLog,
    Label,
    Module,
    ModuleIssue,
    Project,
    ProjectMember,
    State,
)


@pytest.mark.contract
class TestProjectIssueListXlsxExport:
    @staticmethod
    def get_export_url(workspace_slug: str, project_id: str) -> str:
        return f"/api/workspaces/{workspace_slug}/projects/{project_id}/issues/export/xlsx/"

    @staticmethod
    def load_rows(response):
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        return list(worksheet.iter_rows(values_only=True))

    @pytest.fixture
    def project_with_exportable_issues(self, workspace, create_user):
        project = Project.objects.create(
            name="Export Project",
            identifier="XLSX",
            workspace=workspace,
            created_by=create_user,
            updated_by=create_user,
            issue_views_view=True,
            is_time_tracking_enabled=True,
            guest_view_all_features=True,
        )
        ProjectMember.objects.create(
            project=project,
            workspace=workspace,
            member=create_user,
            role=20,
        )
        state = State.objects.create(
            project=project,
            workspace=workspace,
            name="Todo",
            color="#60646C",
            group="unstarted",
            default=True,
            created_by=create_user,
            updated_by=create_user,
        )

        issue_with_april_logs = Issue.objects.create(
            project=project,
            workspace=workspace,
            state=state,
            priority="high",
            name="Issue with April worklogs",
            start_date=date(2026, 4, 5),
            target_date=date(2026, 4, 25),
            created_by=create_user,
            updated_by=create_user,
        )
        issue_with_only_march_log = Issue.objects.create(
            project=project,
            workspace=workspace,
            state=state,
            priority="low",
            name="Issue with March worklog only",
            start_date=date(2026, 3, 10),
            target_date=date(2026, 3, 20),
            created_by=create_user,
            updated_by=create_user,
        )

        IssueWorkLog.objects.bulk_create(
            [
                IssueWorkLog(
                    issue=issue_with_april_logs,
                    user=create_user,
                    hours=1,
                    log_date=date(2026, 3, 31),
                ),
                IssueWorkLog(
                    issue=issue_with_april_logs,
                    user=create_user,
                    hours=1,
                    log_date=date(2026, 4, 10),
                ),
                IssueWorkLog(
                    issue=issue_with_april_logs,
                    user=create_user,
                    hours=1,
                    log_date=date(2026, 4, 20),
                ),
                IssueWorkLog(
                    issue=issue_with_only_march_log,
                    user=create_user,
                    hours=1,
                    log_date=date(2026, 3, 15),
                ),
            ]
        )

        return {
            "project": project,
            "issues": {
                "with_april_logs": issue_with_april_logs,
                "with_only_march_log": issue_with_only_march_log,
            },
        }

    @pytest.mark.django_db
    def test_exports_filtered_project_issue_list_with_summary_and_recalculated_fact(
        self,
        session_client,
        workspace,
        project_with_exportable_issues,
    ):
        project = project_with_exportable_issues["project"]
        issue_with_april_logs = project_with_exportable_issues["issues"]["with_april_logs"]

        response = session_client.post(
            self.get_export_url(workspace.slug, project.id),
            data={
                "layout": "list",
                "rich_filters": {
                    "worklog_log_date__range": "2026-04-01,2026-04-30",
                },
                "applied_filters": {
                    "layout": "list",
                    "priority": "high",
                },
                "columns": [
                    {"key": "name", "label": "Название"},
                    {"key": "key", "label": "ID/key"},
                    {"key": "fact", "label": "Факт"},
                ],
                "filter_summary": ["Период worklog: 2026-04-01 - 2026-04-30"],
            },
            format="json",
        )

        assert response.status_code == status.HTTP_200_OK
        assert response["Content-Disposition"].endswith(".xlsx\"")

        rows = self.load_rows(response)
        assert rows[0][:2] == ("Project", project.name)
        assert rows[2][:2] == ("Filters", "Период worklog: 2026-04-01 - 2026-04-30")
        assert rows[4] == ("Название", "ID/key", "Факт")
        data_rows = rows[5:]
        assert data_rows == [
            (
                issue_with_april_logs.name,
                f"{project.identifier}-{issue_with_april_logs.sequence_id}",
                2,
            )
        ]
        assert len(rows) == 6

    @pytest.mark.django_db
    @pytest.mark.parametrize("layout", ["calendar", "gantt_chart"])
    def test_forces_name_key_and_dates_for_calendar_and_gantt_layouts(
        self,
        session_client,
        workspace,
        project_with_exportable_issues,
        layout,
    ):
        project = project_with_exportable_issues["project"]
        issue_with_april_logs = project_with_exportable_issues["issues"]["with_april_logs"]

        response = session_client.post(
            self.get_export_url(workspace.slug, project.id),
            data={
                "layout": layout,
                "columns": [
                    {"key": "priority", "label": "Priority"},
                ],
                "filter_summary": [],
            },
            format="json",
        )

        assert response.status_code == status.HTTP_200_OK

        rows = self.load_rows(response)
        assert rows[4] == ("Name", "ID", "Start Date", "Due Date")
        assert (
            issue_with_april_logs.name,
            f"{project.identifier}-{issue_with_april_logs.sequence_id}",
            issue_with_april_logs.start_date.strftime("%a, %d %b %Y"),
            issue_with_april_logs.target_date.strftime("%a, %d %b %Y"),
        ) in rows[5:]

    @pytest.mark.django_db
    def test_joins_multi_value_columns_and_keeps_empty_relations_exportable(
        self,
        session_client,
        workspace,
        create_user,
        project_with_exportable_issues,
    ):
        project = project_with_exportable_issues["project"]
        issue_with_april_logs = project_with_exportable_issues["issues"]["with_april_logs"]
        issue_with_only_march_log = project_with_exportable_issues["issues"]["with_only_march_log"]

        label = Label.objects.create(
            name="Finance",
            color="#123456",
            project=project,
            workspace=workspace,
            created_by=create_user,
            updated_by=create_user,
        )
        module = Module.objects.create(
            name="Billing",
            project=project,
            workspace=workspace,
            created_by=create_user,
            updated_by=create_user,
        )
        IssueAssignee.objects.create(
            issue=issue_with_april_logs,
            assignee=create_user,
            project=project,
            workspace=workspace,
            created_by=create_user,
            updated_by=create_user,
        )
        IssueLabel.objects.create(
            issue=issue_with_april_logs,
            label=label,
            project=project,
            workspace=workspace,
            created_by=create_user,
            updated_by=create_user,
        )
        ModuleIssue.objects.create(
            issue=issue_with_april_logs,
            module=module,
            project=project,
            workspace=workspace,
            created_by=create_user,
            updated_by=create_user,
        )

        response = session_client.post(
            self.get_export_url(workspace.slug, project.id),
            data={
                "layout": "list",
                "columns": [
                    {"key": "name", "label": "Название"},
                    {"key": "assignee", "label": "Назначенные"},
                    {"key": "labels", "label": "Метки"},
                    {"key": "modules", "label": "Модули"},
                ],
                "filter_summary": [],
            },
            format="json",
        )

        assert response.status_code == status.HTTP_200_OK

        rows = self.load_rows(response)
        data_rows = rows[5:]
        expected_assignee = f"{create_user.first_name} {create_user.last_name}".strip() or create_user.email
        assert (
            issue_with_april_logs.name,
            f"{project.identifier}-{issue_with_april_logs.sequence_id}",
            expected_assignee,
            "Finance",
            "Billing",
        ) in data_rows
        assert any(
            row[0] == issue_with_only_march_log.name
            and row[1] == f"{project.identifier}-{issue_with_only_march_log.sequence_id}"
            and all(cell in (None, "") for cell in row[2:])
            for row in data_rows
        )
